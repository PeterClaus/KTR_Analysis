import os
import openpyxl
import time
import datetime
import pandas as pd

path1 = "/Users/w0z0341/Desktop/2021_Origin"
path2 = "/Users/w0z0341/Desktop/2021_Temp1"
files = os.listdir(path1)
for f in files:
    if f.endswith(".txt") and f[0].isalpha():
        read_file = pd.read_csv(path1 + "/" + f)
        read_file.to_excel(path1 + "/" + f[:-4] + ".xlsx")

files = os.listdir(path1)
for f in files:
    if f.endswith(".xlsx") and f[0].isalpha():
        wb = openpyxl.load_workbook(path1 + "/" + f)
        ws = wb.active
    else:
        continue
    row = 3
    while row <= 10000:
        date = ws.cell(row, 3).value
        if date is None:
            break
        if ws.cell(row, 4).value != ".TIF":
            ws.delete_rows(row)
            continue
        temp = datetime.datetime.strptime(date, "%m/%d/%Y %H:%M:%S %p")
        sec = time.mktime(temp.timetuple())
        ws.cell(row, 8).value = sec
        startSec = ws.cell(3, 8).value
        ws.cell(row, 9).value = sec - startSec
        row += 1

    wb.save(path2 + "/" + f[:-5] + "_Merge.xlsx")
