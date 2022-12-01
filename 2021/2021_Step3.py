import collections
import os
import openpyxl

path = r"C:\Users\GouldLab\Desktop\2021\2021_Destination"

files = os.listdir(path)
Map = collections.defaultdict(list)
for f in files:
    if f.endswith(".xlsx") and f[0].isalpha():
        wb = openpyxl.load_workbook(path + "/" + f)
        ws = wb.active
        max_row = ws.max_row
        l1, l2 = [], []
        for row in range(1, max_row + 1):
            l1.append(ws.cell(row, 2).value)
            l2.append(ws.cell(row, 5).value)
        if "rep" in f:
            Map[f[:6]].append(l1)
            Map[f[:6]].append(l2)
        else:
            Map[f[:3]].append(l1)
            Map[f[:3]].append(l2)
WB = openpyxl.Workbook()
for K in Map:
    sheet = WB.create_sheet(K)
    WS = WB[K]
    row = 1
    while row <= len(Map[K][0]):
        WS.cell(row, 1).value = Map[K][0][row - 1]
        WS.cell(row, 2).value = Map[K][1][row - 1]
        row += 1
WB.save(path + "/" + "2021_Master.xlsx")