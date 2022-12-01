import collections
import os
import openpyxl

path1 = r"C:\Users\GouldLab\Desktop\2021\Temp1"
path2 = r"C:\Users\GouldLab\Desktop\2021\Temp2"
path3 = r"C:\Users\GouldLab\Desktop\2021\2021_Destination"

files1 = os.listdir(path1)
files2 = os.listdir(path2)

Map = collections.defaultdict(list)

for f in files1:
    if f.endswith(".xlsx") and f[0].isalpha():
        wb1 = openpyxl.load_workbook(path1 + "/" + f)
        ws1 = wb1["Sheet1"]

        for row in range(3, 10000):
            sec = ws1.cell(row, 9).value
            if sec is None:
                break
            if "rep" in f:
                Map[f[:6]].append(sec)
            else:
                Map[f[:3]].append(sec)
        wb1.close()

for K in Map:
    for f in files2:
        if f.startswith(K):
            wb2 = openpyxl.load_workbook(path2 + "/" + f)
            ws2 = wb2.active
            curList = Map[K]
            for i in range(len(curList)):
                ws2.cell(i+2, 2).value = curList[i]
            ws2.cell(1, 5).value = "C/N Ratio"
            for row in range(2, 10000):
                if ws2.cell(row, 6) is None:
                    break
                s = 0
                count = 0
                col = 6


                while col >= 6:
                    if ws2.cell(row, col).value is None:
                        break
                    C = ws2.cell(row, col + 1).value
                    N = ws2.cell(row, col).value
                    try:
                        s += C/N
                        count += 1
                        col += 2
                    except:
                        if int(N) == 0:
                            col += 2
                            continue
                        else:
                            print("Other Errors!!!!!!")
                            col += 2
                            continue
                if count == 0:
                    continue
                ws2.cell(row, 5).value = format(s/count, ".2f")
            if "rep" in f:
                wb2.save(path3 + "/" + f[:6] + "_result.xlsx")
            else:
                wb2.save(path3 + "/" + f[:3] + "_result.xlsx")







