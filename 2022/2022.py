import os
import openpyxl

path1 = r"C:\\Users\\claus\\Desktop\\2022_Origin"
path2 = r"C:\\Users\\claus\\Desktop\\2022_Destination"

files = os.listdir(path1)
data = [("Name", "Average C/N")]
for f in files:
    if f.endswith(".xlsx"):
        wb = openpyxl.load_workbook(path1 + "/" + f)
        ws = wb.active
        row = 2
        s = 0
        count = 0
        while row >= 2:
            if ws.cell(row, 1).value is None:
                break
            C = ws.cell(row + 1, 3).value
            N = ws.cell(row, 3).value
            s += C/N
            count += 1
            row += 2
        if count == 0:
            continue
        data.append((f[:-5], format(s/count, ".2f")))

WB = openpyxl.Workbook()
WS = WB.active
for row in range(1, len(data) + 1):
    WS.cell(row, 1).value = data[row - 1][0]
    WS.cell(row, 2).value = data[row - 1][1]
WB.save(path2 + "/" + "2022_Master.xlsx")